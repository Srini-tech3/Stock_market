import json
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Border, Font, Side
from openpyxl.utils import get_column_letter


class TradeTracker:
    """Export strategy results to Excel workbooks and maintain a journal tracker."""

    JOURNAL_COLUMNS = [
        "Rank",
        "ExpiryDate",
        "Strategy",
        "Option",
        "SellStrike",
        "BuyStrike",
        "SellPremium",
        "BuyPremium",
        "EstimatedPOP",
        "Standalone Funds",
        "Standalone Margin",
        "Profit",
        "Loss",
        "Available Fund",
    ]

    MANUAL_COLUMNS = [
        "Standalone Funds",
        "Standalone Margin",
        "Profit",
        "Loss",
    ]

    def __init__(self, output_dir):
        self.output_dir = Path(output_dir)

    def export_strategy_results(self, bull_put_df, bear_call_df, expiry_date=None, capital=None, file_name=None):
        if expiry_date is not None:
            expiry_str = expiry_date.isoformat() if hasattr(expiry_date, "isoformat") else str(expiry_date)
            if file_name is None:
                file_name = f"strategy_results_{expiry_str}.xlsx"
        else:
            file_name = file_name or "strategy_results.xlsx"

        columns_to_exclude = [
            "WingWidth",
            "SellBid",
            "SellAsk",
            "BuyBid",
            "BuyAsk",
            "SellIV",
            "BuyIV",
            "SellDelta",
            "BuyDelta",
            "SellTheta",
            "BuyTheta",
            "SellVega",
            "BuyVega",
            "SellPOP",
            "BuyPOP",
            "SellSpreadPct",
            "BuySpreadPct",
            "SellOI",
            "BuyOI",
            "SellVolume",
            "BuyVolume",
            "SpreadWidth",
            "NetVega",
            "LiquidityScore",
            "CreditEfficiency",
            "SafetyMargin",
        ]

        bull_put_report = bull_put_df.copy()
        bear_call_report = bear_call_df.copy()

        if expiry_date is not None:
            expiry_str = expiry_date.isoformat() if hasattr(expiry_date, "isoformat") else str(expiry_date)
            bull_put_report["ExpiryDate"] = expiry_str
            bear_call_report["ExpiryDate"] = expiry_str

        bull_put_report["Option"] = bull_put_report["Strategy"].apply(self._option_type_for_strategy)
        bear_call_report["Option"] = bear_call_report["Strategy"].apply(self._option_type_for_strategy)

        if "Rank" in bull_put_report.columns:
            bull_put_report = bull_put_report.sort_values("Rank")
        if "Rank" in bear_call_report.columns:
            bear_call_report = bear_call_report.sort_values("Rank")

        bull_put_report = bull_put_report.drop(
            columns=[col for col in columns_to_exclude if col in bull_put_report.columns],
            errors="ignore",
        )
        bear_call_report = bear_call_report.drop(
            columns=[col for col in columns_to_exclude if col in bear_call_report.columns],
            errors="ignore",
        )

        if "Rank" in bull_put_report.columns and "ExpiryDate" in bull_put_report.columns:
            cols = list(bull_put_report.columns)
            expiry_index = cols.index("ExpiryDate")
            cols.insert(cols.index("Rank") + 1, cols.pop(expiry_index))
            bull_put_report = bull_put_report[cols]

        if "Rank" in bear_call_report.columns and "ExpiryDate" in bear_call_report.columns:
            cols = list(bear_call_report.columns)
            expiry_index = cols.index("ExpiryDate")
            cols.insert(cols.index("Rank") + 1, cols.pop(expiry_index))
            bear_call_report = bear_call_report[cols]

        if "Strategy" in bull_put_report.columns and "Option" in bull_put_report.columns:
            cols = list(bull_put_report.columns)
            cols.insert(cols.index("Strategy") + 1, cols.pop(cols.index("Option")))
            bull_put_report = bull_put_report[cols]

        if "Strategy" in bear_call_report.columns and "Option" in bear_call_report.columns:
            cols = list(bear_call_report.columns)
            cols.insert(cols.index("Strategy") + 1, cols.pop(cols.index("Option")))
            bear_call_report = bear_call_report[cols]

        self.output_dir.mkdir(parents=True, exist_ok=True)
        output_file = self.output_dir / file_name

        with pd.ExcelWriter(output_file, engine="openpyxl") as writer:
            bull_put_report.to_excel(writer, sheet_name="BullPut", index=False)
            bear_call_report.to_excel(writer, sheet_name="BearCall", index=False)

        self._write_strategy_result_json(bull_put_report, bear_call_report, expiry_date)
        self._format_workbook(output_file)
        return output_file

    def export_journal_tracker(self, bull_put_df, bear_call_df, expiry_date=None, capital=300000, file_name=None):
        expiry_str = expiry_date.isoformat() if hasattr(expiry_date, "isoformat") else str(expiry_date) if expiry_date is not None else None

        if file_name is None:
            file_name = "journal_tracker.xlsx"

        self.output_dir.mkdir(parents=True, exist_ok=True)
        output_file = self.output_dir / file_name

        bull_journal = self._build_journal_sheet(bull_put_df, expiry_str)
        bear_journal = self._build_journal_sheet(bear_call_df, expiry_str)
        journal_df = pd.concat([bull_journal, bear_journal], ignore_index=True)

        if output_file.exists():
            existing = pd.read_excel(output_file, sheet_name="Journal")
            journal_df = self._merge_journal_rows(existing, journal_df)

        with pd.ExcelWriter(output_file, engine="openpyxl") as writer:
            journal_df.to_excel(writer, sheet_name="Journal", index=False)

        self._add_available_fund_formulas(output_file, capital)
        self._format_workbook(output_file)
        return output_file

    def _build_journal_sheet(self, df, expiry_str):
        report = df.copy() if df is not None else pd.DataFrame()

        report = report.copy()
        if "Strategy" not in report.columns:
            report["Strategy"] = None

        if "Rank" in report.columns:
            report = report.sort_values("Rank")
        else:
            report = report.copy()

        report["ExpiryDate"] = expiry_str

        for col in [
            "Rank",
            "ExpiryDate",
            "Strategy",
            "SellStrike",
            "BuyStrike",
            "SellPremium",
            "BuyPremium",
            "EstimatedPOP",
        ]:
            if col not in report.columns:
                report[col] = None

        report["Option"] = report["Strategy"].apply(self._option_type_for_strategy)

        report = report[[
            "Rank",
            "ExpiryDate",
            "Strategy",
            "Option",
            "SellStrike",
            "BuyStrike",
            "SellPremium",
            "BuyPremium",
            "EstimatedPOP",
        ]]

        for col in self.MANUAL_COLUMNS:
            report[col] = None

        report["Available Fund"] = None
        return report[self.JOURNAL_COLUMNS]

    def _merge_journal_rows(self, existing, current):
        identity_columns = [
            "Rank",
            "ExpiryDate",
            "Strategy",
            "Option",
            "SellStrike",
            "BuyStrike",
            "SellPremium",
            "BuyPremium",
            "EstimatedPOP",
        ]

        existing = existing.copy()
        current = current.copy()

        existing = existing.reindex(columns=self.JOURNAL_COLUMNS)
        current = current.reindex(columns=self.JOURNAL_COLUMNS)

        if existing.empty and current.empty:
            return pd.DataFrame(columns=self.JOURNAL_COLUMNS)

        merged = pd.concat([existing, current], ignore_index=True)
        merged = merged.drop_duplicates(subset=identity_columns, keep="last")

        for _, row in existing.iterrows():
            if row[identity_columns].isna().any():
                continue

            mask = pd.Series(True, index=merged.index)
            for col in identity_columns:
                mask &= merged[col].astype(object).eq(row[col])

            if mask.any():
                for col in self.MANUAL_COLUMNS:
                    if pd.notna(row[col]):
                        merged.loc[mask, col] = row[col]

        return merged.reindex(columns=self.JOURNAL_COLUMNS)

    def _option_type_for_strategy(self, strategy):
        strategy_value = str(strategy).strip().lower()
        if strategy_value == "bull put":
            return "PE"
        if strategy_value == "bear call":
            return "CE"
        return None

    def _write_strategy_result_json(self, bull_put_report, bear_call_report, expiry_date=None):
        payload = {
            "expiry_date": self._stringify_expiry(expiry_date),
            "bull_put": self._dataframe_to_records(bull_put_report),
            "bear_call": self._dataframe_to_records(bear_call_report),
        }

        result_path = self.output_dir / "result.json"
        with result_path.open("w", encoding="utf-8") as handle:
            json.dump(payload, handle, indent=2, ensure_ascii=False)

    def _dataframe_to_records(self, dataframe):
        if dataframe is None or dataframe.empty:
            return []
        return dataframe.where(pd.notna(dataframe), None).to_dict(orient="records")

    def _stringify_expiry(self, expiry_date):
        if expiry_date is None:
            return None
        return expiry_date.isoformat() if hasattr(expiry_date, "isoformat") else str(expiry_date)

    def _format_workbook(self, output_file):
        workbook = load_workbook(output_file)
        font = Font(size=16)
        border = Border(
            left=Side(border_style="thin", color="000000"),
            right=Side(border_style="thin", color="000000"),
            top=Side(border_style="thin", color="000000"),
            bottom=Side(border_style="thin", color="000000"),
        )

        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            max_widths = {}

            for row in sheet.iter_rows():
                for cell in row:
                    cell.font = font
                    cell.border = border
                    value = cell.value
                    if value is None:
                        length = 0
                    else:
                        length = len(str(value))
                    max_widths[cell.column_letter] = max(max_widths.get(cell.column_letter, 0), length)

            for column_letter, width in max_widths.items():
                sheet.column_dimensions[column_letter].width = max(width + 3, 10)

        workbook.save(output_file)

    def _preserve_manual_columns(self, report, existing_sheet):
        if existing_sheet is None:
            return report

        existing = existing_sheet.copy()
        if "Rank" not in existing.columns or "Strategy" not in existing.columns:
            return report

        existing = existing[[col for col in ["Rank", "Strategy"] + self.MANUAL_COLUMNS if col in existing.columns]]
        report = report.merge(existing, on=["Rank", "Strategy"], how="left", suffixes=("", "_existing"))

        for col in self.MANUAL_COLUMNS:
            existing_col = f"{col}_existing"
            if existing_col in report.columns:
                report[col] = report[existing_col].combine_first(report[col])
                report = report.drop(columns=[existing_col])

        return report[self.JOURNAL_COLUMNS]

    def _add_available_fund_formulas(self, output_file, capital):
        workbook = load_workbook(output_file)

        sheet_name = "Journal"
        if sheet_name not in workbook.sheetnames:
            workbook.save(output_file)
            return

        sheet = workbook[sheet_name]
        headers = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]

        if "Available Fund" not in headers or "Standalone Funds" not in headers:
            workbook.save(output_file)
            return

        avail_col = get_column_letter(headers.index("Available Fund") + 1)
        funds_col = get_column_letter(headers.index("Standalone Funds") + 1)

        for row in range(2, sheet.max_row + 1):
            if row == 2:
                formula = f"={capital}-{funds_col}{row}"
            else:
                formula = f"={avail_col}{row - 1}-{funds_col}{row}"
            sheet[f"{avail_col}{row}"] = formula

        workbook.save(output_file)
