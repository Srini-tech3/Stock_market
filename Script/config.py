from pathlib import Path
from datetime import date, timedelta
from openpyxl import load_workbook
import pandas as pd
import re


class ConfigManager:
    """Manage project configuration paths and option-chain metadata."""

    def __init__(
        self,
        input_dir: Path | str | None = None,
        output_dir: Path | str | None = None,
        config_file: Path | str | None = None,
    ):
        self.input_dir = Path(input_dir or "E:/StockMarket_Project/OptionAnalyzer/Input")
        self.output_dir = Path(output_dir or "E:/StockMarket_Project/OptionAnalyzer/Output")
        self.config_file = Path(config_file or self.input_dir / "Config.xlsx")

    def load_config(self):
        df = pd.read_excel(self.config_file, sheet_name="CONFIG")
        return dict(zip(df["Parameter"], df["Value"]))

    def _extract_expiry_from_name(self, filename):
        match = re.search(r"_(\d{4}-\d{2}-\d{2})_option_chain", filename, re.IGNORECASE)

        if match:
            return pd.to_datetime(match.group(1)).date()

        return None

    def find_option_chain(self, expiry_date=None):
        files = list(self.input_dir.glob("*.csv"))

        if len(files) == 0:
            raise FileNotFoundError("No CSV file found inside Input folder.")

        if expiry_date is not None:
            matching = [f for f in files if self._extract_expiry_from_name(f.name) == expiry_date]
            if not matching:
                available = [self._extract_expiry_from_name(f.name) for f in files]
                available_str = ", ".join(str(e) for e in available if e is not None)
                raise FileNotFoundError(
                    f"No option chain CSV found for expiry {expiry_date}. Available expiries: {available_str}"
                )
            if len(matching) > 1:
                raise Exception(f"Multiple option chain CSV files found for expiry {expiry_date}.")
            return matching[0]

        if len(files) == 1:
            return files[0]

        raise Exception("Multiple CSV files found. Pass --expiry YYYY-MM-DD to select one.")

    def extract_expiry(self, csv_file):
        expiry = self._extract_expiry_from_name(csv_file.name)

        if expiry is None:
            raise Exception("Unable to detect expiry date from filename.")

        return expiry

    def get_trading_date(self):
        today = date.today()

        if today.weekday() == 5:
            today += timedelta(days=2)

        elif today.weekday() == 6:
            today += timedelta(days=1)

        return today

    def get_dte(self, expiry, trading_day):
        return (expiry - trading_day).days

    def update_config_values(self, values: dict):
        wb = load_workbook(self.config_file)
        ws = wb["CONFIG"]

        existing = {}

        for row in range(2, ws.max_row + 1):
            parameter = ws.cell(row=row, column=1).value
            existing[parameter] = row

        for key, value in values.items():
            if key in existing:
                row = existing[key]
                ws.cell(row=row, column=2).value = value
            else:
                row = ws.max_row + 1
                ws.cell(row=row, column=1).value = key
                ws.cell(row=row, column=2).value = value

        wb.save(self.config_file)